from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

ARCHIVO = "usuarios.xlsx"

# 🔹 crear archivo si no existe
if not os.path.exists(ARCHIVO):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["ID", "Nombre"])
    wb.save(ARCHIVO)

def guardar_en_excel(nombre):
    wb = load_workbook(ARCHIVO)
    ws = wb.active

    # calcular ID
    nuevo_id = ws.max_row

    ws.append([nuevo_id, nombre])
    wb.save(ARCHIVO)

@app.route("/")
def inicio():
    return render_template("index.html")

@app.route("/guardar_usuario", methods=["POST"])
def guardar_usuario():
    data = request.json
    nombre = data["nombre"]

    guardar_en_excel(nombre)

    return jsonify({"mensaje": "usuario guardado en excel"})

@app.route("/mensaje", methods=["POST"])
def mensaje():
    data = request.json
    texto = data["mensaje"]

    if texto.lower() == "hola":
        respuesta = "Hola"
    else:
        respuesta = "No entiendo el mensaje"

    return jsonify({"respuesta": respuesta})

if __name__ == "__main__":
    app.run(debug=True)